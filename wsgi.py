#!/usr/bin/env python
# coding: utf-8

# #### Week 3 - Functions

# In[ ]:


# Import Library default with python installation
import os
import sys


# In[ ]:


# line magic in jupyter notebook
# get_ipython().system('dir ')


# In[ ]:


def my_main(input_filename):
    print('***Start of my_main function***')
    
    # Get the current working directory 
    curr_dir = os.getcwd()
    print(curr_dir)

    # Get the list of files in the directory
    file_list = os.listdir()
    print(file_list)

    # Search thru the list for the input file name
    input_file_found = False
    inputfilegood=False

    for afile in file_list:
        if afile == input_filename:
            input_file_found = True
            print('Input File found', input_filename)
            break
        #End If
    #End For

    # Is the File extension .txt
    if input_filename.lower().endswith('.txt'):
        inputfilegood=True
        print('Input File extension good', input_filename)
    else:
        print('***ERROR*** Input file extension NOT good')
#       sys.exit('***Exiting***')
        return(-1, -1, -1,[],"***ERROR*** Input file extension NOT good") 
    #End If

    # Do we continue ?
    if  input_file_found == True and inputfilegood == True:
        print('Input file is found and extension good')
    else:
        print('***ERROR*** Input file NOT found OR extension NOT good')
#       sys.exit('***Exiting***')
        return(-1, -1, -1,[],'***ERROR*** Input file NOT found OR extension NOT good')
    #End If

    # Ready to open Input file
    # Mode read 
    # File Handle is the File descriptor - path , mode, name 
    try:
        input_file_obj = open(input_filename,'r')
        print('*** INFO***',input_filename,' Input file opened successfully')
    except:
        print('*** ERROR***',input_filename,' Input file CANNOT OPEN')
#       sys.exit('\n***Exiting***')
        return(-1, -1, -1,[],'*** ERROR***',input_filename,' Input file CANNOT OPEN')

    # Audit the file 
    print('\n ***Starting Audit***')

    # Create a list of the word lengths we want to track 0 to 16
    word_length = []

    max = 17
    for i in range(0, max, 1):
        word_length.append(i)
    #End For    

    print(word_length)    

    # Frequency table will keep track of the frequency of the word lengths
    freq_table = []
    freq_table = 17*[0]        # initialize outside loop with 17 zeros
    print(freq_table)

    # Initialize line totals
    linecount = 0
    wordcount =0
    charcount = 0
    nowscharcount = 0

    # Initialize File totals
    totlinecount = 0
    totwordcount = 0
    totcharcount = 0
    totnowscharcount = 0

    # Initialize Lists for Openpyxl
    line_Content = []
    line_Count = []
    word_Count =[]
    char_Count = []
    nowschar_Count=[]

    # Initialze output list for flask
    output_list = []

    # Read the input file into a list

    for myline in input_file_obj:
        print(myline)

        linecount = linecount + 1
        word_list = myline.split()
        print(word_list)
        wordcount = len(word_list)

        charcount = len(myline)

        nowscharcount = 0
        for myword in word_list:
            nowscharcount = nowscharcount + len(myword)
            if len(myword) >= max - 1:
                freq_table[max-1] = freq_table[max-1]+1
            else:
                freq_table[len(myword)] = freq_table[len(myword)] + 1
            #End If
        #End for

        print(linecount, wordcount, charcount , nowscharcount)

        final_line = myline + "," + str(linecount) + "," + str(wordcount) + "," + str(charcount) + "," + str(nowscharcount)
        output_list.append(final_line)

        totlinecount = linecount
        totwordcount = totwordcount + wordcount
        totcharcount = totcharcount + charcount
        totnowscharcount =  totnowscharcount + nowscharcount

        line_Content.append(myline)
        line_Count.append(linecount)
        word_Count.append(wordcount)
        char_Count.append(charcount)
        nowschar_Count.append(nowscharcount)

    #End For

    print(word_length)
    print(freq_table)

    # Print File totals
    print(totlinecount)
    print(totwordcount)
    print(totcharcount)
    print(totnowscharcount)

    

    #### Week 2 - Openpyxl

    # pip install library NOT default with python installation
    import openpyxl

    # Create workbook obj
    wb = openpyxl.Workbook()
    xlsx_filename = input_filename.replace(".txt", ".xlsx")

    # Create sheet1
    ws1 = wb.active
    ws1.title = input_filename.replace(".txt", "")

    # column headers for worksheet 1
    ws1.cell(row=1,column=1).value = "LINE_CONTENT"
    ws1.cell(row=1,column=2).value = "LINE_COUNT"
    ws1.cell(row=1,column=3).value = "WORD_COUNT"
    ws1.cell(row=1,column=4).value = "CHAR_COUNT"
    ws1.cell(row=1,column=5).value = "NOWS_CHAR_COUNT"

    # Data for Sheet1
    for index in range(0,len(line_Content),1):
        ws1.cell(row=index+2,column=1).value = line_Content[index]
        ws1.cell(row=index+2,column=2).value = line_Count[index]
        ws1.cell(row=index+2,column=3).value = word_Count[index]
        ws1.cell(row=index+2,column=4).value = char_Count[index]
        ws1.cell(row=index+2,column=5).value = nowschar_Count[index]
    #End For   

    # Create sheet2
    ws2 = wb.create_sheet("WORD_ANALYSIS")

    # Column headers for worksheet 2
    ws2.cell(row=1,column=1).value = "WORD_LENGTH"
    ws2.cell(row=1,column=2).value = "WORD_COUNT"

    # Data for Sheet2
    for index in range(0, max, 1):
        ws2.cell(row=index+2,column=1).value = "WORD_LENGTH_" + str(index)
        ws2.cell(row=index+2,column=2).value = freq_table[index]
    #End For    

    # Designate rows/columns as the labels and Data for the charts
    labels = openpyxl.chart.Reference(ws2, min_col=1, min_row=2, max_row=max+1)
    data = openpyxl.chart.Reference(ws2, min_col=2, min_row=1, max_row=max+1)

    # Draw the pie chart for the data in the ANALYSIS worksheet
    pie = openpyxl.chart.PieChart()
    pie.title  = 'WORD_ANALYSIS'
    pie.add_data(data,titles_from_data=True)
    pie.set_categories(labels)
    ws2.add_chart(pie, "F5")

    # Draw the Bar chart for the data in the BarChartAnalysis worksheet
    bar =  openpyxl.chart.BarChart()
    bar.shape = 4
    bar.style = 10
    bar.title = 'WORD_ANALYSIS'
    bar.y_axis.title = 'WORD_COUNT'
    bar.x_axis.title = 'WORD_LENGTH'
    bar.add_data(data,titles_from_data=True)
    bar.set_categories(labels)
    ws2.add_chart(bar, 'F30')

    # Save the workbook
    wb.save(xlsx_filename)

    #Close files after finishing audit
    print('*** INFO*** Closing files')
    input_file_obj.close()
    wb.close()
    
    return(totlinecount, totwordcount, totcharcount,output_list,xlsx_filename,"" )

# end of function my_main()

# In[ ]:


# # Main Program for Standalone App
# # If __name__ = __main__ ,program is running standalone
# if __name__ == "__main__":
#     print("Python script is run standalone\n")
        
# #   Get the name from the user
#     name = input()   

# #   Call fx Main program
#     tlc, twc, tcc, output_list, errstr = my_main(name)

#     print(tlc)
#     print(twc)
#     print(tcc)
#     print(output_list)

# else:
#     # __name__ will have the name of the module that imported this script
#     print("Python script was imported")     
# #End Main program


# #### Week 4 - Flask

# In[ ]:


# Import libraries
# flask will look for the templates folder
from flask import Flask
from flask import render_template 
from flask import request
from flask import send_file


# In[ ]:

from werkzeug.utils import secure_filename
# Import werkzeug to run your app as a web application
# from werkzeug.serving import run_simple

OUTFILE = ""
# In[ ]:


# Main Program for Web App
# If __name__ = __main__ ,program is running standalone
if __name__ == "__main__":
    
    # if __main__, file was main file run standalone
    print ("\nPython script is run standalone\n")
    print("Python special variable __name__ =", __name__)

     # Run the flask app in jupyter notebook needs run_simple 
     # Run the flask app in python script in OpenShift needs application.run
    application.run(host='0.0.0.0', debug=True)

else:
    # __name__ will have the name of the module that imported this script
    print("Python script was imported")
    
    ##############################################
    # flask setup
    ##############################################
    
    # Instantiate the Flask object 
    application = Flask(__name__)

    # home displays the selectform.html
    @application.route('/', methods=['GET'])
    def welcomepage():
        return('<h1><center>Text Audit Welcome Page</h1><br>')
    # end of function home1()
 
    # home displays the selectform.html
    @application.route('/home', methods=['GET'])
    def home():
        return render_template('selectform.html')
    # end of function home()

    # submit on the selectform.html will conduct the audit
    @application.route('/audit', methods=['POST'])
    def audit():

        global OUTFILE
        # file_obj = request.files.get('txtdata')
        # print("Type of the file is :", type(file_obj))
        # name = file_obj.filename
        # print(name)

        #############################################
        # get file info to upload file to container
        #
        print("*** request.files['txtdata']")
        file_obj = request.files['txtdata']
        print("*** Type of the file_obj is :", type(file_obj))
        filename = secure_filename(file_obj.filename)
        file_obj.save(os.path.join(os.getcwd(), filename))
        print("*** inout file name is:", os.path.join(os.getcwd(), filename))
        #############################################

        # Call fx Main program
        tlc, twc, tcc, output_list, outfile, errstr = my_main(filename)
        OUTFILE = outfile
        print("OUTFILE = ", OUTFILE)

        if len(errstr):
            return render_template('selectform.html', errstr= errstr)
        else:
            # return render_template('template.html',
            #                     my_string=filename,
            #                     line_count=tlc,
            #                     total_word_count = twc,
            #                     total_char_count=tcc,
            #                     my_list=output_list,)
            # return(send_file(os.path.join(os.getcwd(), outfile), as_attachment=True))
            return render_template('results.html',
                                my_string=filename,
                                line_count=tlc,
                                total_word_count = twc,
                                total_char_count=tcc,
                                my_list=output_list,)
    # end of audit()


    @application.route('/return-file/')
    def return_file():
        try:
            print("OUTFILE = ", OUTFILE)
            return(send_file(os.path.join(os.getcwd(), OUTFILE), as_attachment=True))
        except Exception as e:
            return str(e)

# end of if __name__ == "__main__": ... else

#End of Audit_Text program

